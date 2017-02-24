from shutil import copyfile
from openpyxl import Workbook
from openpyxl import load_workbook

TABLE_HEIGHT = 4
SHEET_OFFSET = 1

copyfile("template.xlsx", "output.xlsx")
i = 0
workbook = load_workbook("output.xlsx")
worksheet = workbook.active

with open("input.txt", "r") as input_file, open("output.txt", "w") as output_file:

	for _ in range(1000):

		line = input_file.readline()

		if line.startswith("PP"):

			registers = input_file.readline().split()
			registers = registers[:5] + registers[5].split(":") + registers[6:]
			registers[7] = registers[7] + " " + registers[8]
			registers = registers[1:-1]
			command = input_file.readline().split()[-2:]
			command = " ".join(command)

			worksheet.cell(
				column=1,
				row=i*TABLE_HEIGHT + SHEET_OFFSET + 1,
				value=command,
			)

			for j, register in enumerate(registers):

				worksheet.cell(
					column=j + 1,
					row=i*TABLE_HEIGHT + SHEET_OFFSET + 2 + 1,
					value=register,
				)

			i += 1

workbook.save(filename="output.xlsx")